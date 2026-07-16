import importlib.util
import sys
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


# MCP 以 stdio 脚本方式启动，测试也按该入口加载。
_MODULE_PATH = Path(__file__).parents[4] / "package/yuxi/agents/mcp/buildin/document_exporter.py"
_SPEC = importlib.util.spec_from_file_location("document_exporter_test", _MODULE_PATH)
assert _SPEC and _SPEC.loader
document_exporter = importlib.util.module_from_spec(_SPEC)
sys.modules[_SPEC.name] = document_exporter
_SPEC.loader.exec_module(document_exporter)


async def test_mcp_exposes_document_export_tools():
    tools = await document_exporter.mcp.list_tools()

    assert {tool.name for tool in tools} == {"generate_docx", "generate_pdf", "generate_xlsx"}


def test_generate_docx_creates_readable_document(tmp_path, monkeypatch):
    monkeypatch.setattr(document_exporter, "_ARTIFACT_DIR", tmp_path)

    result = document_exporter.generate_docx("会议纪要", "# 结论\n按期交付", "项目会议")

    document = Document(result["artifact_path"])
    assert [paragraph.text for paragraph in document.paragraphs] == ["项目会议", "结论", "按期交付"]
    assert result["filename"] == "会议纪要.docx"


async def test_generate_pdf_uses_office_converter(tmp_path, monkeypatch):
    monkeypatch.setattr(document_exporter, "_ARTIFACT_DIR", tmp_path)

    async def fake_converter(filename: str, content: bytes) -> bytes:
        assert filename == "document.docx"
        assert content[:2] == b"PK"
        return b"%PDF-1.7\n"

    monkeypatch.setattr(document_exporter, "convert_office_to_pdf", fake_converter)
    result = await document_exporter.generate_pdf("报告", "正文")

    assert Path(result["artifact_path"]).read_bytes() == b"%PDF-1.7\n"
    assert result["filename"] == "报告.pdf"


def test_generate_xlsx_creates_requested_sheets(tmp_path, monkeypatch):
    monkeypatch.setattr(document_exporter, "_ARTIFACT_DIR", tmp_path)

    result = document_exporter.generate_xlsx("统计", {"汇总": [["项目", "数量"], ["A", 2]]})

    workbook = load_workbook(result["artifact_path"], data_only=True)
    assert workbook.sheetnames == ["汇总"]
    assert list(workbook["汇总"].values) == [("项目", "数量"), ("A", 2)]
    assert result["filename"] == "统计.xlsx"
